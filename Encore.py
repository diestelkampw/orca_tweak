#! python3
import openpyxl
from openpyxl import Workbook
import os
import sys
import datetime


def add_headers(template, new_quote):
    for row in template.iter_rows(min_row=1, max_col=20, max_row=1, values_only=True):
        new_quote.append(row)


def get_quote_num(sheet):
    return sheet['E2'].value


def update_manufacturer(name):
    match name:
        case 'Planar Systems':
            manufacturer = 'PLANAR'
        case 'Chief':
            manufacturer = 'CHIEF MANUFACTURING'
        case 'Soundcontrol':
            manufacturer = 'SOUND CONTROL TECHNOLOGIES'
        case 'Shure':
            manufacturer = 'SHURE INCORPORATED'
        case 'Netgear':
            manufacturer = 'NETGEAR, INC.'
        case 'Middle Atlantic':
            manufacturer = 'MIDDLE ATLANTIC PRODUCTS INC'
        case 'ADB':
            manufacturer = 'GENERAL CABLE'
        case 'SharpNEC':
            manufacturer = 'NEC COMPUTERS'
        case 'Extron Electronics':
            manufacturer = 'ELECTRONICS/RGB SYSTEMS'
        case 'Blackbox':
            manufacturer = 'Black Box'
        case 'Biamp Systems':
            manufacturer = 'Biamp'
        case 'Xtreme Power Conversion':
            manufacturer = 'EXTREME'
        case 'Chatsworth':
            manufacturer = 'Chatsworth Products Inc'
        case 'Samsung':
            manufacturer = 'Samsung America, Inc'
        case 'Startech':
            manufacturer = 'Startech Computer'
        case _:
            manufacturer = name
    return manufacturer


def add_data(encore, new_quote, quote_num):
    vendor = 'Encore Technologies'
    for row in encore.iter_rows(min_row=7, max_col=7, max_row=encore.max_row, values_only=True):
        print(row[1])
        if row[1] is not None:
            part_number = row[2]
            description = row[3]
            price = row[4]
            quantity = row[5]
            manufacturer = update_manufacturer(row[1])
            new_row = [part_number, description, '', price, '', quantity, manufacturer, vendor,
                       '', '', '', '', '', '', quote_num]
            print(new_row)
            new_quote.append(new_row)
        else:
            continue


def main():
    input_file = 'Encore.xlsx'
    output_file = 'converted_file_' + str(datetime.datetime.now()) + '.xlsx'
    output_file = output_file.replace(":", "-")

    output_file_path = os.path.join("C:/Users/diestelw/OneDrive - WWT/Scripts/OrcaTweak/",
                                    output_file)
    orca_template_path = os.path.join("C:/Users/diestelw/OneDrive - WWT/Scripts/OrcaTweak/",
                                      "Quote Import Template.xlsx")

    try:
        encore_wb = openpyxl.load_workbook(input_file)
    except PermissionError:
        sys.exit("File access denied.")
    try:
        orca_wb = openpyxl.load_workbook(orca_template_path)
    except PermissionError:
        sys.exit("File access denied.")

    new_quote_wb = Workbook()

    encore_sheet = encore_wb.active
    orca_sheet = orca_wb.active
    new_quote_sheet = new_quote_wb.active

    for merge in list(encore_sheet.merged_cells):
        encore_sheet.unmerge_cells(range_string=str(merge))

    add_headers(orca_sheet, new_quote_sheet)

    quote_num = get_quote_num(encore_sheet)

    add_data(encore_sheet, new_quote_sheet, quote_num)

    new_quote_wb.save(output_file_path)
    new_quote_wb.close()
    encore_wb.close()
    orca_wb.close()


if __name__ == '__main__':
    main()
