#! python3
import openpyxl
from openpyxl import Workbook
import os
import sys
import datetime
import re
from html2excel import ExcelParser


def convert_html(input_file, output_file):
    parser = ExcelParser(input_file)
    parser.to_excel(output_file)


def add_headers(template, new_quote):
    for row in template.iter_rows(min_row=1, max_col=20, max_row=1, values_only=True):
        new_quote.append(row)


def get_quote_num(dell):
    for row in dell.iter_rows(min_row=1, max_col=3, max_row=dell.max_row, values_only=True):
        if row[1] == "Quote #:":
            return row[2]


def add_data(dell, new_quote, quote_num):
    in_row = False
    grab_data = False
    first_row = True
    for row in dell.iter_rows(min_row=1, max_col=4, max_row=dell.max_row, values_only=True):
        if in_row:
            if grab_data:
                if row[0] is None:
                    in_row = False
                    grab_data = False
                    first_row = True
                    new_row = ['', '', '', '', '', '', '', '']
                    new_quote.append(new_row)
                    continue
                else:
                    data = re.match(r"(.*)\((.+)\)", row[0])
                    part_number = data[2]
                    description = data[1].strip()
                    if first_row:
                        new_row = [part_number, description, '', price, '', row[3], 'Dell', 'Dell Federal Systems L.P.',
                                   '', '', '', '', '', '', quote_num]
                        new_quote.append(new_row)
                        first_row = False
                    else:
                        new_row = [part_number, description, '', 0, '', row[3], 'Dell', 'Dell Federal Systems L.P.',
                                   '', '', '', '', '', '', quote_num]
                        new_quote.append(new_row)
            else:
                if 'Quantity' in row[3]:
                    grab_data = True
                    continue
                else:
                    continue
        elif row[0] is None:
            continue
        elif "GROUP:" in row[0]:
            in_row = True
            data = re.findall(r"\:\W+([\d\,\.]+)", row[0])
            group_num = data[1]
            price = float(data[2].replace(',', ''))


def main():
    input_file = 'text_file.html'
    output_file = 'converted_file_' + str(datetime.datetime.now()) + '.xlsx'
    output_file = output_file.replace(":", "-")
    convert_html(input_file, output_file)

    dell_file_path = os.path.join("C:/Users/diestelw/OneDrive - WWT/Scripts/OrcaTweek/",
                                  output_file)
    orca_template_path = os.path.join("C:/Users/diestelw/OneDrive - WWT/Scripts/OrcaTweek/",
                                      "Quote Import Template.xlsx")

    try:
        dell_wb = openpyxl.load_workbook(dell_file_path)
    except PermissionError:
        sys.exit("File access denied.")
    try:
        orca_wb = openpyxl.load_workbook(orca_template_path)
    except PermissionError:
        sys.exit("File access denied.")

    new_quote_wb = Workbook()

    dell_sheet = dell_wb.active
    orca_sheet = orca_wb.active
    new_quote_sheet = new_quote_wb.active

    for merge in list(dell_sheet.merged_cells):
        dell_sheet.unmerge_cells(range_string=str(merge))

    add_headers(orca_sheet, new_quote_sheet)

    quote_num = get_quote_num(dell_sheet)

    add_data(dell_sheet, new_quote_sheet, quote_num)

    dell_save_path = os.path.join("C:/Users/diestelw/OneDrive - WWT/Scripts/OrcaTweek/",
                                  "DellQuote_" + quote_num + ".xlsx")

    new_quote_wb.save(dell_save_path)
    new_quote_wb.close()
    dell_wb.save(dell_file_path)
    dell_wb.close()
    orca_wb.close()


if __name__ == '__main__':
    main()
