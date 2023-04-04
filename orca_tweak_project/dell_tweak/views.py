from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import FileSystemStorage
import openpyxl
from openpyxl import Workbook
import os
from os.path import exists
import sys
import re
from html2excel import ExcelParser

# Create your views here.


def home(request):
    return render(request, 'dell_tweak/home.html')


def dell(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)

        output_file = 'converted_file_' + myfile.name
        output_file = output_file.replace(":", "-")
        output_file = output_file.replace(".html", ".xlsx")
        output_path = os.path.join(settings.MEDIA_ROOT, output_file)

        if exists(output_path):
            os.remove(output_path)

        input_path = os.path.join(settings.MEDIA_ROOT, myfile.name)
        convert_html(input_path, output_path)

        orca_template_path = os.path.join(settings.BASE_DIR, "Quote Import Template.xlsx")

        try:
            dell_wb = openpyxl.load_workbook(output_path)
        except PermissionError:
            sys.exit("File access denied.")
        try:
            orca_wb = openpyxl.load_workbook(orca_template_path)
        except PermissionError:
            sys.exit("File access denied.")

        new_quote_wb = Workbook()

        dell_sheet = dell_wb.active
        orca_sheet = orca_wb.active
        new_quote_sheet = new_quote_wb.active

        for merge in list(dell_sheet.merged_cells):
            dell_sheet.unmerge_cells(range_string=str(merge))

        add_headers(orca_sheet, new_quote_sheet)

        quote_num = get_quote_num(dell_sheet, 'dell')

        add_dell_data(dell_sheet, new_quote_sheet, quote_num)
        dell_filename = "DellQuote_" + quote_num + ".xlsx"
        dell_save_path = os.path.join(settings.MEDIA_ROOT, dell_filename)
        dell_file_url = settings.MEDIA_URL + dell_filename

        new_quote_wb.save(dell_save_path)
        new_quote_wb.close()
        dell_wb.close()
        orca_wb.close()
        os.remove(input_path)
        os.remove(output_path)

        return render(request, 'dell_tweak/dell.html', {'dell_file_url': dell_file_url})
    return render(request, 'dell_tweak/dell.html')


def encore(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)

        input_file = myfile.name
        input_path = os.path.join(settings.MEDIA_ROOT, input_file)
        orca_template_path = os.path.join(settings.BASE_DIR, "Quote Import Template.xlsx")

        try:
            encore_wb = openpyxl.load_workbook(input_path)
        except PermissionError:
            sys.exit("File access denied.")
        try:
            orca_wb = openpyxl.load_workbook(orca_template_path)
        except PermissionError:
            sys.exit("File access denied.")

        new_quote_wb = Workbook()
        encore_sheet = encore_wb.active
        orca_sheet = orca_wb.active
        new_quote_sheet = new_quote_wb.active

        for merge in list(encore_sheet.merged_cells):
            encore_sheet.unmerge_cells(range_string=str(merge))

        add_headers(orca_sheet, new_quote_sheet)
        quote_num = get_quote_num(encore_sheet, 'encore')
        add_encore_data(encore_sheet, new_quote_sheet, quote_num)

        encore_filename = "EncoreQuote_" + quote_num + ".xlsx"
        encore_save_path = os.path.join(settings.MEDIA_ROOT, encore_filename)
        encore_file_url = settings.MEDIA_URL + encore_filename

        new_quote_wb.save(encore_save_path)
        new_quote_wb.close()
        encore_wb.close()
        orca_wb.close()
        os.remove(input_path)

        return render(request, 'dell_tweak/encore.html', {'encore_file_url': encore_file_url})
    return render(request, 'dell_tweak/encore.html')


def convert_html(input_file, output_file):
    parser = ExcelParser(input_file)
    parser.to_excel(output_file)


def add_headers(template, new_quote):  # re-usable
    for row in template.iter_rows(min_row=1, max_col=20, max_row=1, values_only=True):
        new_quote.append(row)


def get_quote_num(sheet, partner):
    if partner == 'dell':
        for row in sheet.iter_rows(min_row=1, max_col=3, max_row=sheet.max_row, values_only=True):
            if row[1] == "Quote #:":
                return row[2]
    elif partner == 'encore':
        return sheet['E2'].value


def add_dell_data(dell, new_quote, quote_num):
    in_row = False
    grab_data = False
    first_row = True
    for row in dell.iter_rows(min_row=1, max_col=4, max_row=dell.max_row, values_only=True):
        if in_row:
            if grab_data:
                if row[0] is None:
                    in_row = False
                    grab_data = False
                    first_row = True
                    new_row = ['', '', '', '', '', '', '', '']
                    new_quote.append(new_row)
                    continue
                else:
                    data = re.match(r"(.*)\((.+)\)", row[0])
                    part_number = data[2]
                    description = data[1].strip()
                    if first_row:
                        new_row = [part_number, description, '', price, '', row[3], 'Dell', 'Dell Federal Systems L.P.',
                                   '', '', '', '', '', '', quote_num]
                        new_quote.append(new_row)
                        first_row = False
                    else:
                        new_row = [part_number, description, '', 0, '', row[3], 'Dell', 'Dell Federal Systems L.P.',
                                   '', '', '', '', '', '', quote_num]
                        new_quote.append(new_row)
            else:
                if 'Quantity' in row[3]:
                    grab_data = True
                    continue
                else:
                    continue
        elif row[0] is None:
            continue
        elif "GROUP:" in row[0]:
            in_row = True
            data = re.findall(r"\:\W+([\d\,\.]+)", row[0])
            group_num = data[1]
            price = float(data[2].replace(',', ''))


def add_encore_data(encore, new_quote, quote_num):
    vendor = 'Encore Technologies'
    for row in encore.iter_rows(min_row=7, max_col=7, max_row=encore.max_row, values_only=True):
        print(row[1])
        if row[1] is not None:
            part_number = row[2]
            description = row[3]
            price = row[4]
            quantity = row[5]
            if row[1] == 'Planar Systems':
                manufacturer = 'PLANAR'
            elif row[1] == 'Chief':
                manufacturer = 'CHIEF MANUFACTURING'
            elif row[1] == 'Soundcontrol':
                manufacturer = 'SOUND CONTROL TECHNOLOGIES'
            elif row[1] == 'Shure':
                manufacturer = 'SHURE INCORPORATED'
            elif row[1] == 'Netgear':
                manufacturer = 'NETGEAR, INC.'
            elif row[1] == 'Middle Atlantic':
                manufacturer = 'MIDDLE ATLANTIC PRODUCTS INC'
            elif row[1] == 'ADB':
                manufacturer = 'GENERAL CABLE'
            else:
                manufacturer = row[1]
            new_row = [part_number, description, '', price, '', quantity, manufacturer, vendor,
                       '', '', '', '', '', '', quote_num]
            print(new_row)
            new_quote.append(new_row)
        else:
            continue
